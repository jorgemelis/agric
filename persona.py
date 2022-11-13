import os, sys, shutil, re
import openpyxl
import pandas as pd

from lib import info as inf


def personalidad(beneficiario, info):
    ret = "PF"
    pj_contains = info["pj_contains"]
    pj_ends_with = info["pj_ends_with"]
    pj_has_word = info["pj_has_word"]
    aer_contains = info["aer_contains"]
    aer_ends_with = info["aer_ends_with"]

    if any(x.lower() in beneficiario.lower() for x in aer_contains):
        ret = "AER"
    if any(beneficiario.lower().endswith(x.lower()) for x in aer_ends_with):
        ret = "AER"
    if any(x.lower() in beneficiario.lower() for x in pj_contains):
        ret = "PJ"
    if any(beneficiario.lower().endswith(x.lower()) for x in pj_ends_with):
        ret = "PJ"
    if any(re.search(rf"\b{x}\b", beneficiario) for x in pj_has_word):
        ret = "PJ"
    return ret


info = inf.get_info()
pj_ends_with = info["pj_ends_with"]
for s in pj_ends_with:
    print("pj_ends_with:", s)

# print(type(info), info)
# sys.exit()

# Fichero limpio (sin tipo de persona)
xlsin = "/Users/jorge/git/agric/data/2021_beneficiarios_original.xlsx"
xlspath = "/Users/jorge/git/agric/data/2021_beneficiarios.xlsx"
shutil.copy(xlsin, xlspath)

# relleno de personalidad
book = openpyxl.load_workbook(xlspath)
sheet = book["Sheet1"]
sheet.cell(row=1, column=4).value = "Tipo"
for nrow in range(2, sheet.max_row + 1):
    beneficiario = sheet.cell(row=nrow, column=1).value
    pj = personalidad(beneficiario, info)
    sheet.cell(row=nrow, column=4).value = pj
    # print(nrow, beneficiario, pj)
book.save(xlspath)


xlspath = "/Users/jorge/git/agric/data/2021_beneficiarios.xlsx"
dfin = pd.read_excel(xlspath)
# print(dfin)
dfpj = dfin[dfin["Tipo"] == "PJ"]
dfaer = dfin[dfin["Tipo"] == "AER"]
dfpf = dfin[dfin["Tipo"] == "PF"]
# string = "herederos"
# dftmp = dfin[dfin["Beneficiario"].str.contains(string, case=False)]

xlsout = "/Users/jorge/git/agric/out/2021_personas.xlsx"
with pd.ExcelWriter(xlsout, engine="xlsxwriter") as writer:
    dfpj.to_excel(writer, sheet_name="PJ", index=False)
    dfaer.to_excel(writer, sheet_name="AER", index=False)
    dfpf.to_excel(writer, sheet_name="PF", index=False)
    for string in info["list_strings"]:
        dftmp = dfin[dfin["Beneficiario"].str.contains(string, case=False)]
        dftmp.to_excel(writer, sheet_name=string, index=False)

print("done")
